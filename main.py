import customtkinter as ctk
import openpyxl as opxl
from openpyxl import Workbook
import string
import pyperclip
from tkinter import messagebox
import os
from PIL import Image
from random import randint as rd

def gerar_senha():
    try:
        qt = int(entrada_qt_numeros_senhas.get())
        if qt < 1:
            raise ValueError
    except ValueError:
        messagebox.showerror("Erro", "Digite um número maior que 0")
        return

    caracteres = string.ascii_letters + string.digits + string.punctuation
    senha = ''.join([caracteres[rd(0, len(caracteres)-1)]for _ in range(qt)])
    senha_criada.configure(text=senha)

def copiar_senha():
    senha = senha_criada.cget("text")
    if senha:
        pyperclip.copy(senha)
        messagebox.showinfo("Senha Copiada")

def salvar_na_planilha():
    senha = senha_criada.cget("text")
    finalidade = entrada_finalidade.get()

    if not senha:
        messagebox.showinfo("Aviso", "Nenhuma senha gerada.")
        return

    if not finalidade:
        messagebox.showinfo("Aviso", "Digite a finalidade da senha.")
        return

    arquivo = "senhasrandompass.xlsx"

    try:
        if os.path.exists(arquivo):
            planilha = opxl.load_workbook(arquivo)
            aba = planilha.active
        else:
            planilha = opxl.Workbook()
            aba = planilha.active
            aba.append(["Finalidade", "Senha"])
        
        aba.append([finalidade ,senha])
        planilha.save(arquivo)
        messagebox.showinfo("Sucesso", "Senha salva com Sucesso")

    except ValueError as e:
        messagebox.showerror("Erro", f"Erro ao salvar: {str(e)}")

def abrir_planilha():
    caminho = "senhasrandompass.xlsx"
    if os.path.exists(caminho):
        os.startfile(caminho)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Senhas"
        ws.append(["Finalidade", "Senha"])
        wb.save(caminho)
        messagebox.showinfo("Planilha Criada com Sucesso")
 
    try:
        os.startfile(caminho)
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível abrir a planilha:\n{e}")

#GERAÇÃO DA JANELA
janela = ctk.CTk()
janela._set_appearance_mode("system")
janela.title("RandomPass - Senhas seguras aqui")
janela.geometry("700x500")
janela.resizable(height="False", width="False")

#FRAME QUE COMPORTA TODOS OS OUTROS BOTÕES E FRAMES
frame_pai = ctk.CTkFrame(janela, width=100, height=100, fg_color="#0b004d")
frame_pai.pack(pady=20)

titulo = ctk.CTkLabel(frame_pai, text="RandomPass", font=("arial black", 32), fg_color="transparent", bg_color="transparent", 
                      corner_radius=0, text_color= "#00bf63")
titulo.pack(padx=200, pady= 50)


#ONDE O USUÁRIO DIZ PARA QUÊ É A SENHA E A QUANTIDADE DE CARACTERES QUE ELE QUER
frame_formacao_senha = ctk.CTkFrame(frame_pai)
frame_formacao_senha.pack()
entrada_finalidade = ctk.CTkEntry(frame_formacao_senha, placeholder_text="Sistema/site para senha", width=205, height=20,
                                        fg_color="transparent", bg_color="transparent", corner_radius=80)
entrada_finalidade.pack(pady=(0, 5))

entrada_qt_numeros_senhas= ctk.CTkEntry(frame_formacao_senha, placeholder_text="Digite a quantidade de caracteres", width=205, 
                                        height=20,fg_color="transparent", bg_color="transparent", corner_radius=80)
entrada_qt_numeros_senhas.pack(padx=10, pady=(0, 5), side="left")

#ONDE FICARÃO A SENHA GERADA E P BOTÃOD E COPIAR A SENHA
frame_senha = ctk.CTkFrame(frame_pai, width= 200, height=30)
frame_senha.pack(pady=10)
frame_senha_gerada = ctk.CTkFrame(frame_senha, fg_color="#e2e2e2", width=200, height=30, bg_color="transparent",
                                  corner_radius=0)
frame_senha_gerada.pack_propagate(False)
frame_senha_gerada.pack(pady=10, padx=30)
botao_gerar_senha = ctk.CTkButton(frame_senha, text="Gerar Senha", width=80,height=30, fg_color="#00bf63", 
                                  corner_radius=80, command=gerar_senha)
botao_gerar_senha.pack(padx=10)
senha_criada = ctk.CTkLabel(frame_senha_gerada, text="", font=("arial", 14), text_color="black")
senha_criada.pack(expand=True)

#ONDE FICARÃO OS BOTÕES DE APOIO PARA INTERAÇÃO COM A SENHA
frame_botoes = ctk.CTkFrame(frame_pai, bg_color="transparent", fg_color="transparent")
frame_botoes.pack(pady=10)

botao_copiar_senha = ctk.CTkButton(frame_botoes, text="Copiar", width=80, height=30, fg_color="#00bf63",
                                    corner_radius=80, command=copiar_senha)
botao_copiar_senha.pack(padx=10)

botao_salvar_na_planilha = ctk.CTkButton(frame_botoes, text="Salvar na Planilha", width=80, height=30, 
                                         fg_color="#00bf63", corner_radius=80, command=salvar_na_planilha)
botao_salvar_na_planilha.pack(padx=10, pady=10, side="left")

botao_abrir_planilha = ctk.CTkButton(frame_botoes, text="Abrir Planilha", width=80, height=30, 
                                    fg_color="#00bf63", corner_radius=80, command=abrir_planilha)
botao_abrir_planilha.pack(padx=10, side="left")

janela.mainloop() 